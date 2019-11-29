
import openpyxl as op
import psycopg2
from psycopg2.extras import DictCursor #чтобы выводился список вместо кортежа


class postgree_connect:

	''' При создании экземпляра прописыватьимя бд для подключения'''
	''' имя DB прописывать строкой '''

	def __init__(self, db_name):
		self.conn = psycopg2.connect(dbname = db_name, user = 'postgres', 
                        password = '2378951', host = 'localhost')
		self.cursor = self.conn.cursor(cursor_factory = DictCursor)
 
	def insert(self, table_name, values):
		''' Для инсерта прописывать имя таблицы строкой'''
		''' Values - передавать кортежем '''

		#self.cursor.execute("INSERT INTO {} VALUES ('ABC', '32213121') ".format(table_name)) 
		self.cursor.execute("INSERT INTO {} VALUES {}".format(table_name, values)) 
		self.conn.commit() #сохраняем изменения в БД

	def show_table(self, table_name):
		''' Для показа всех данных прописывать имя таблицы строкой'''

		self.cursor.execute("SELECT * FROM {}".format(table_name)) 
		for row in self.cursor:
			print(row)

	def close(self):		
		''' всегда закрывать каретку и БД'''

		self.cursor.close() #закрываем каретку/курсор
		self.conn.close() #закрываем БД


	def __str__(self):
		return 'class Object postgree DB'





class open_xlsx:

	def __init__(self, file_name, sheet_number):
		''' Путь к файлу передавать строкой '''
		''' Номер листа для копирования передавать целочисленным'''
		self.wb = op.load_workbook(filename = file_name)
		self.sheets = self.wb.sheetnames #загружаем список листов


		self.ws = self.wb[self.sheets[sheet_number-1]] #передаем номер листа
		''' Здесь обращение происходит по индексу, поэтому -> минус один'''

		self.A_column = self.ws['A'] 
		self.B_column = self.ws['B'] 
		self.C_column = self.ws['C'] 
		self.D_column = self.ws['D'] 
		self.E_column = self.ws['E'] 
		self.F_column = self.ws['F'] 
		self.G_column = self.ws['G'] 
		self.H_column = self.ws['H'] 
		self.I_column = self.ws['I'] 
		self.J_column = self.ws['J'] 
		self.K_column = self.ws['K'] 
		self.L_column = self.ws['L'] 
		self.M_column = self.ws['M'] 
		self.N_column = self.ws['N'] 
		self.O_column = self.ws['O'] 
		self.P_column = self.ws['P'] 
		self.Q_column = self.ws['Q'] 
		self.R_column = self.ws['R'] 
		self.S_column = self.ws['S'] 
		self.T_column = self.ws['T'] 
		self.U_column = self.ws['U'] 
		self.V_column = self.ws['V'] 
		self.W_column = self.ws['W'] 
		self.X_column = self.ws['X'] 
		self.Y_column = self.ws['Y'] 
		self.Z_column = self.ws['Z'] 
		self.AA_column = self.ws['AA'] 
		self.AB_column = self.ws['AB'] 
		self.AC_column = self.ws['AC'] 
		self.AD_column = self.ws['AD'] 
		self.AE_column = self.ws['AE'] 
		self.AF_column = self.ws['AF'] 
		self.AG_column = self.ws['AG'][1:]

		

	def take_the_values(self, name_column_end):
		''' Имя колонки писать строкой, этот аргумент определяет конец считывания '''
		test_cort = []
		start_index = 1		
		for column in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 
					   'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',
					   'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG'     				   ]:
						
			test_cort.append(self.ws[column][1].value)
			if column == name_column_end:
				break

		print(test_cort)
		
		'''
			if column == name_column_end:
				break
			else:
				for column_value in self.ws[column][1:]:
					print(column_value.value)
		'''

#a = postgree_connect('test_db')
#a.show_table('test_table')
#a.close()

op_xl_1 = open_xlsx('d:/Users/A/Desktop/тест.xlsx', 1)
op_xl_1.take_the_values('D')